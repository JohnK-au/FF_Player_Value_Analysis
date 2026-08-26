"""Pre-auction league workbook: all 8 rosters + cap/roster flexibility.

    .venv/bin/python -m src.reports.auction_workbook

Reads the freshly-rebuilt player_pricing_2026.csv (the full per-team join, keyed
on espn_id) + the sheet's authoritative cap figures, and writes an Excel workbook
to data/processed/reports/league_rosters_2026.xlsx (gitignored -- league data).

Tabs: Summary (8 teams side by side, cap + roster flexibility), one per team
(full roster + contracts + value + pricing + a KEEP/DROP/TAG/EXTEND rec), and
Validation (cap reconcile + roster counts + flagged data anomalies).

Built to be iterated -- one function per tab. The FA/auction-targets tab is a
deliberate follow-up.
"""
from __future__ import annotations

import datetime as dt
import difflib
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import CAP_TOTAL, PROCESSED_DIR, TEAMS, UPCOMING_SEASON

MY_TEAM = "Kerr"
POS_ORDER = ["QB", "RB", "WR", "TE"]
VET_SLOTS = 28  # 14 starters + 14 bench veteran contract slots

# --- auction strategy (Kerr) -- edit these as the plan evolves --------------
# Personal max bids (cap units) in the user's priority order. These OVERRIDE the
# model's fair value where the user's context is better (e.g. Adams on a 1yr
# rental -- the model's fair is an age-discounted DYNASTY price).
TARGET_BIDS = [
    {"player": "Davante Adams",       "your_max": 120,
     "note": "ANCHOR. Best producer of the WR targets; 1yr rental so ignore the dynasty age-discount. Push it."},
    {"player": "Blake Corum",         "your_max": 45,
     "note": "Most-wanted RB. At model fair -- stay disciplined."},
    {"player": "Keaton Mitchell",     "your_max": 35,
     "note": "2nd RB want. Cheap upside flier."},
    {"player": "Dalton Kincaid",      "your_max": 40,
     "note": "Only if cheap. Pairs with cutting Andrews (opens TE1). Walk away above 40."},
    {"player": "Michael Wilson",      "your_max": 20,
     "note": "Only if <20. Upgrade over the WRs you're cutting (prod 38 > McMillan/Tucker/TeSlaa/Boutte)."},
    {"player": "Travis Etienne",      "your_max": 75,
     "note": "Low interest. Only at a discount."},
]

# Cut candidates on Kerr, in the user's stated priority order.
CUT_LIST = ["Tre Tucker", "Jalen McMillan", "Mark Andrews", "Isaac TeSlaa", "Kayshon Boutte"]

CAP_RESERVE = 75      # keep for in-season flexibility (user: leave 50-100)
NEW_CUT_RATE = 0.50   # dead cap on NEW cuts (rules going-forward; 0.20 is legacy)
PRICING_CSV = PROCESSED_DIR / "player_pricing_2026.csv"
OUT = PROCESSED_DIR / "reports" / "league_rosters_2026.xlsx"

POS_OVERRIDES = PROCESSED_DIR / "rookie_position_overrides.csv"

TEAM_COLS = ["player", "position_group", "type", "nfl_team", "roster_status",
             "age", "salary_2026", "years_2026", "dynasty_total_salary",
             "dynasty_value", "on_field_value", "replacement_dv",
             "fair_value_2026", "surplus_2026", "fair_value_dynasty",
             "surplus_dynasty", "recommendation"]


def _roster_type(row: pd.Series) -> str:
    """K-IDP-HC (roster filler) / SKILL (valued skill player) / ROOKIE (skill
    prospect with no model value yet -- incl. the hand-assigned rookie positions).
    Keyed on value presence so an overridden rookie position stays ROOKIE."""
    if row["position_group"] in ("K/P", "IDP", "HC"):
        return "K-IDP-HC"
    return "SKILL" if pd.notna(row.get("dynasty_value")) else "ROOKIE"

# --- styling constants ---
HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
KERR_FILL = PatternFill("solid", fgColor="FFF2CC")   # highlight the user's team
GOOD_FILL = PatternFill("solid", fgColor="D6EAD6")    # bargain (surplus << 0)
BAD_FILL = PatternFill("solid", fgColor="F4CCCC")     # overpaid (surplus >> 0)
FLAG_FILL = PatternFill("solid", fgColor="FCE5CD")    # validation flag


def _f(v, default=0.0):
    """NaN-safe float."""
    return float(v) if pd.notna(v) else default


def _recommend(row: pd.Series) -> str:
    """DROP/TAG/EXTEND/KEEP -- mirrors src/app/_lib.py::recommend_action."""
    surplus = row.get("surplus_2026")
    if pd.isna(surplus):
        return "-"
    surplus, salary = float(surplus), _f(row.get("salary_2026"))
    years = int(_f(row.get("years_2026")))
    dv, fair = _f(row.get("dynasty_value")), _f(row.get("fair_value_2026"))
    if salary >= 40 and surplus > 60:
        return "DROP"
    if years == 1 and dv >= 70 and surplus <= -60 and fair >= 100:
        return "TAG"
    if years == 2 and surplus <= -50 and dv >= 55:
        return "EXTEND"
    return "KEEP"


# ------------------------------------------------------------------ data
# Value columns pulled from the pricing CSV onto the roster ground truth.
_VALUE_COLS = ["years_2026", "dynasty_total_salary",
               "dynasty_value", "on_field_value", "replacement_dv",
               "fair_value_2026", "surplus_2026", "fair_value_dynasty",
               "surplus_dynasty"]


def _tagged_rows(existing: set, cw: pd.DataFrame) -> pd.DataFrame:
    """Franchise-tagged players as roster rows: 1yr @ tag salary, status
    'tagged', value columns joined from pricing, current team/position from the
    2026 crosswalk. The sheet keeps tags in a separate TAG section, so
    player_salaries_2026 omits them. `existing` guards against double-counting a
    tag already in the active section; `cw` = 2026 crosswalk (player/_cur_pos/nfl_team)."""
    from src.data.cap import parse_tags
    from src.data.players import POSITION_GROUP

    tags = parse_tags()
    tags = tags[~tags["player"].isin(existing)].copy()
    if tags.empty:
        return tags
    price = pd.read_csv(PRICING_CSV)
    vcols = [c for c in _VALUE_COLS if c in price.columns]
    price = price[["player", "espn_id", "age", "position_group", *vcols]].drop_duplicates("player")
    t = tags.rename(columns={"salary": "salary_2026"}).merge(price, on="player", how="left")
    t["years_2026"] = 1
    t["roster_status"] = "tagged"
    t["dynasty_total_salary"] = t["salary_2026"]  # a tag is a one-year deal
    t = t.merge(cw, on="player", how="left")
    grp = t["_cur_pos"].map(POSITION_GROUP)
    need = t["position_group"].isna() | t["position_group"].eq("Other")
    t.loc[need, "position_group"] = grp[need]
    return t.drop(columns=[c for c in ["_cur_pos", "league_year", "tag_year"] if c in t.columns])


def load_roster() -> pd.DataFrame:
    """Every rostered player from the sheet (ground truth for ownership) with
    value columns joined on espn_id. Rookies/unmatched players appear with their
    team + salary but blank value columns -- framework drops them (no position to
    score), so they'd otherwise be missing entirely."""
    if not PRICING_CSV.exists():
        raise SystemExit(f"{PRICING_CSV} missing -- run the refresh first.")
    from src.data.cap import player_salaries_2026

    sal = player_salaries_2026().rename(columns={"source": "roster_status"})
    price = pd.read_csv(PRICING_CSV)
    vcols = [c for c in _VALUE_COLS if c in price.columns]
    pv = price.loc[price["espn_id"].notna(), ["espn_id", *vcols]].copy()
    # merge on a float64 key -- espn_id is nullable Int64 and pandas<2 can't
    # factorize NA keys (the unmatched rookies); astype float64 turns NA into a
    # real NaN that merges safely (unmatched rows just get blank value columns).
    sal["_id"] = sal["espn_id"].astype("float64")
    pv["_id"] = pv["espn_id"].astype("float64")
    roster = sal.merge(pv.drop(columns="espn_id"), on="_id", how="left").drop(columns="_id")

    # Current NFL team + position from the 2026 ESPN crosswalk. Built off the
    # live 2026 rosters, so it captures offseason moves the 2025-based value
    # crosswalk misses (AJ Brown PHI->NE) AND matches the deep rookies it
    # couldn't (100% coverage). DISPLAY only -- the value pipeline still keys on
    # the 2025 crosswalk/espn_id, so Team-component multipliers are unchanged.
    from src.data.players import POSITION_GROUP

    cw = pd.read_csv(PROCESSED_DIR / "player_crosswalk_2026.csv")[["player", "position", "pro_team"]]
    cw = cw.drop_duplicates("player").rename(columns={"position": "_cur_pos", "pro_team": "nfl_team"})
    roster = roster.merge(cw, on="player", how="left")
    # fill missing/"Other" position_group (the rookies) from the current crosswalk
    cur_grp = roster["_cur_pos"].map(POSITION_GROUP)
    need = roster["position_group"].isna() | roster["position_group"].eq("Other")
    roster.loc[need, "position_group"] = cur_grp[need].fillna(roster.loc[need, "position_group"])
    roster = roster.drop(columns="_cur_pos")

    # optional hand-override (wins over the crosswalk) -- now a safety net, since
    # the 2026 pull already covers every rookie; kept for any future manual fix.
    if POS_OVERRIDES.exists():
        ov = pd.read_csv(POS_OVERRIDES).set_index("player")["position_group"]
        m = roster["player"].isin(ov.index)
        roster.loc[m, "position_group"] = roster.loc[m, "player"].map(ov)

    # Fold in franchise-tagged players -- a real 1yr roster slot the sheet keeps
    # in its own TAG section, so player_salaries_2026 omits them: without this
    # they vanish from their team's roster AND wrongly show as FA-pool available.
    tagged = _tagged_rows(set(roster["player"]), cw)
    if len(tagged):
        roster = pd.concat([roster, tagged], ignore_index=True)

    roster["type"] = roster.apply(_roster_type, axis=1)
    roster["recommendation"] = roster.apply(_recommend, axis=1)
    return roster


def cap_figures() -> pd.DataFrame:
    """Sheet's authoritative CAP USED / DEAD CAP / CAP SPACE per team for 2026."""
    from src.data.cap import parse_cap_summary
    caps = parse_cap_summary()
    caps = caps[caps["season"] == UPCOMING_SEASON].set_index("team")
    return caps[["cap_used", "dead_cap", "cap_space"]]


def reconcile_table() -> pd.DataFrame:
    from src.data.cap import reconcile
    return reconcile(UPCOMING_SEASON).set_index("team")


# ------------------------------------------------------------------ tabs
def build_summary(roster: pd.DataFrame, caps: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for team in TEAMS:
        r = roster[roster["team"] == team]
        vet = int(r["roster_status"].isin(["active", "extension", "tagged"]).sum())
        cap = caps.loc[team] if team in caps.index else None
        rows.append({
            "team": team,
            "players": len(r),
            "vet": vet,
            "rookie": int((r["roster_status"] == "rookie").sum()),
            "ps": int((r["roster_status"] == "practice_squad").sum()),
            "open_vet_slots": VET_SLOTS - vet,
            "cap_used": _f(cap["cap_used"]) if cap is not None else np.nan,
            "cap_space": _f(cap["cap_space"]) if cap is not None else np.nan,
            "dead_cap": _f(cap["dead_cap"]) if cap is not None else np.nan,
            "tot_dynasty_value": round(r["dynasty_value"].sum(), 1),
            "net_surplus_2026": round(r["surplus_2026"].sum(), 1),
            "net_surplus_dynasty": round(r["surplus_dynasty"].sum(), 1),
            "n_bargains": int((r["surplus_2026"] < -30).sum()),
            "n_overpaid": int((r["surplus_2026"] > 30).sum()),
        })
    return pd.DataFrame(rows)


def build_team(roster: pd.DataFrame, team: str) -> pd.DataFrame:
    """Full roster for `team` (ground truth from the sheet). Skill positions
    first sorted by surplus; rookies/unmatched (position 'Other', blank values)
    listed last."""
    r = roster[roster["team"] == team].copy()
    pos_rank = {p: i for i, p in enumerate(POS_ORDER)}
    r["_pos"] = r["position_group"].map(pos_rank).fillna(len(POS_ORDER))
    r = r.sort_values(["_pos", "surplus_2026"], na_position="last")
    cols = [c for c in TEAM_COLS if c in r.columns]
    return r[cols].round(1)


def build_validation(roster: pd.DataFrame, recon: pd.DataFrame,
                     notes: list[str]) -> tuple[pd.DataFrame, list[str]]:
    rows = []
    for team in TEAMS:
        rr = recon.loc[team] if team in recon.index else None
        n = int((roster["team"] == team).sum())
        resid = _f(rr["resid"]) if rr is not None else np.nan
        rows.append({
            "team": team, "roster_count": n,
            "cap_used_sheet": _f(rr["sheet"]) if rr is not None else np.nan,
            "recon": _f(rr["recon"]) if rr is not None else np.nan,
            "resid (sheet-recon)": round(resid, 1),
            "flag": "CHECK" if abs(resid) > 15 else "ok",
        })
    return pd.DataFrame(rows), notes


# ------------------------------------------------------------------ auction plan
def _priced_universe() -> pd.DataFrame:
    """Full pricing table joined to current (2026) NFL team. Joins on espn_id
    (via the 2026 ESPN cache) so it covers the FA pool too -- the contract-only
    crosswalk misses free agents."""
    price = pd.read_csv(PRICING_CSV)
    teams = pd.read_csv(PROCESSED_DIR / "espn_teams_2026.csv")
    price["_id"] = price["espn_id"].astype("float64")
    teams["_id"] = teams["espn_id"].astype("float64")
    price = price.merge(teams[["_id", "pro_team"]].drop_duplicates("_id"), on="_id", how="left")
    return price.drop(columns="_id").rename(columns={"pro_team": "nfl_team"})


def _find_player(name: str, pool: list[str]) -> str | None:
    if name in pool:
        return name
    m = difflib.get_close_matches(name, pool, n=1, cutoff=0.55)
    return m[0] if m else None


def build_bid_plan(univ: pd.DataFrame) -> pd.DataFrame:
    """The user's targets + personal max bids, annotated with the model's read.
    Blank RESULT / price_paid columns are for live use during the auction."""
    pool = univ["player"].tolist()
    rows = []
    for i, t in enumerate(TARGET_BIDS, 1):
        m = _find_player(t["player"], pool)
        r = univ[univ["player"] == m].iloc[0] if m is not None else None
        fair = round(_f(r["fair_value_2026"]), 1) if r is not None else np.nan
        rows.append({
            "pri": i, "player": t["player"],
            "pos": r["position_group"] if r is not None else None,
            "nfl_team": r["nfl_team"] if r is not None else None,
            "age": r["age"] if r is not None else np.nan,
            "prod_score": round(_f(r["production_value"]), 0) if r is not None else np.nan,
            "model_fair": fair, "your_max": t["your_max"],
            "premium_vs_model": round(t["your_max"] - fair, 1) if pd.notna(fair) else np.nan,
            "note": t["note"], "RESULT": "", "price_paid": "",
        })
    return pd.DataFrame(rows)


def build_cut_plan(roster: pd.DataFrame) -> pd.DataFrame:
    """Kerr cut candidates with the cap each frees. New-cut dead cap is 50%/yr,
    so current-year relief = half the salary; amnesty frees the full salary."""
    kerr = roster[roster["team"] == MY_TEAM]
    rows = []
    for i, name in enumerate(CUT_LIST, 1):
        hit = kerr[kerr["player"] == name]
        if hit.empty:
            rows.append({"pri": i, "player": name, "pos": "NOT on Kerr roster"})
            continue
        r = hit.iloc[0]
        sal = _f(r["salary_2026"])
        rows.append({
            "pri": i, "player": name, "pos": r["position_group"],
            "nfl_team": r.get("nfl_team"), "salary_2026": round(sal, 1),
            "years": int(_f(r["years_2026"])),
            "model_fair": round(_f(r["fair_value_2026"]), 1),
            "surplus_2026": round(_f(r["surplus_2026"]), 1),
            "cap_freed_cut50": round(sal * (1 - NEW_CUT_RATE), 1),
            "cap_freed_amnesty": round(sal, 1),
        })
    return pd.DataFrame(rows)


def build_budget(caps: pd.DataFrame, cut_plan: pd.DataFrame,
                 bid_plan: pd.DataFrame) -> pd.DataFrame:
    """Kerr's spendable auction budget under the user's spend/reserve philosophy."""
    cap_space = _f(caps.loc[MY_TEAM, "cap_space"]) if MY_TEAM in caps.index else 0.0
    cuts50 = _f(cut_plan["cap_freed_cut50"].sum()) if "cap_freed_cut50" in cut_plan else 0.0
    if "salary_2026" in cut_plan and len(cut_plan):
        big = _f(cut_plan["salary_2026"].max())
        amnesty_extra = big * NEW_CUT_RATE  # full salary - the 50% relief already counted
    else:
        amnesty_extra = 0.0
    planned = _f(bid_plan["your_max"].sum())
    spend_norm = cap_space + cuts50 - CAP_RESERVE
    rows = [
        ("Cap space now (Kerr)", round(cap_space, 1)),
        ("+ Cuts freed (all, normal 50%)", round(cuts50, 1)),
        ("+ Amnesty upgrade on biggest cut (extra)", round(amnesty_extra, 1)),
        ("- In-season reserve", float(-CAP_RESERVE)),
        ("Spendable (normal cuts)", round(spend_norm, 1)),
        ("Spendable (if amnesty biggest)", round(spend_norm + amnesty_extra, 1)),
        ("Planned max bids (all targets)", round(planned, 1)),
        ("Slack if you win ALL at max", round(spend_norm - planned, 1)),
    ]
    return pd.DataFrame(rows, columns=["item", "cap_units"])


def build_fa_pool(univ: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    """Available FAs (unrostered + untagged) ranked by model fair within
    position -- the pivot list if a target gets bid past your max. Targets marked ★."""
    from src.data.cap import parse_tags
    tagged = set(parse_tags()["player"])
    target_names = {_find_player(t["player"], univ["player"].tolist()) for t in TARGET_BIDS}
    fa = univ[(univ["roster_status"] == "fa") & (~univ["player"].isin(tagged))].copy()
    fa = fa[fa["position_group"].isin(POS_ORDER)]
    fa["tgt"] = np.where(fa["player"].isin(target_names), "★", "")
    fa = fa.sort_values(["position_group", "fair_value_2026"], ascending=[True, False])
    # top-N per position, but ALWAYS keep the user's targets even if lower-ranked
    top = fa.groupby("position_group", group_keys=False).head(top_n)
    fa = pd.concat([top, fa[fa["tgt"] == "★"]]).drop_duplicates("player")
    fa = fa.sort_values(["position_group", "fair_value_2026"], ascending=[True, False])
    out = fa[["position_group", "tgt", "player", "nfl_team", "age",
              "production_value", "dynasty_value", "fair_value_2026"]]
    return out.rename(columns={"position_group": "pos", "production_value": "prod_score",
                               "dynasty_value": "DV", "fair_value_2026": "model_fair"}).round(1)


# ------------------------------------------------------------------ styling
def _style_sheet(ws, highlight_team_col: bool = False):
    # header row
    for cell in ws[1]:
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    # column widths from content
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 8), 34)


def _color_surplus(ws, headers: list[str]):
    for name in ("surplus_2026", "surplus_dynasty", "net_surplus_2026"):
        if name not in headers:
            continue
        ci = headers.index(name) + 1
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=ci)
            if isinstance(c.value, (int, float)):
                if c.value < -30:
                    c.fill = GOOD_FILL
                elif c.value > 30:
                    c.fill = BAD_FILL


def main() -> None:
    roster = load_roster()
    caps = cap_figures()
    recon = reconcile_table()

    notes = [
        f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} from a full refresh "
        f"(sheet + ESPN + value rebuild), 2026 base season.",
        "Team 'Couc' was renamed 'Paik' (ownership change).",
        "Contract Extensions tab is empty this cycle (extensions happen AFTER "
        "the FA auction) -- expected.",
        "Cap figures are the SHEET's authoritative CAP USED / SPACE; 'resid' is "
        "our reconstruction gap (5/8 teams exact; Drew/Will/Silv small edges).",
        "2 fuzzy mis-matches blocked (Chris Bell, Omar Cooper Jr) + 12 rookies "
        "unmatched to nflverse -> those players show blank value metrics.",
        "Sheet years-remaining quirks: 'Rams HC' and Terry McLaurin have unreadable "
        "contract-year formulas on the sheet.",
    ]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    summary = build_summary(roster, caps)
    validation, _ = build_validation(roster, recon, notes)
    univ = _priced_universe()
    bid_plan = build_bid_plan(univ)
    cut_plan = build_cut_plan(roster)
    budget = build_budget(caps, cut_plan, bid_plan)
    fa_pool = build_fa_pool(univ)

    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="Summary", index=False)
        # auction-actionable tabs, up front
        bid_plan.to_excel(xw, sheet_name="Targets", index=False)
        cut_plan.to_excel(xw, sheet_name="Cut_Plan", index=False)
        ws = xw.sheets["Cut_Plan"]                      # budget block below the cut table
        bstart = len(cut_plan) + 3
        ws.cell(row=bstart, column=1, value="BUDGET (Kerr)").font = Font(bold=True)
        for j, (item, val) in enumerate(budget.itertuples(index=False), 1):
            ws.cell(row=bstart + j, column=1, value=item)
            ws.cell(row=bstart + j, column=2, value=val)
        fa_pool.to_excel(xw, sheet_name="FA_Pool", index=False)
        for team in TEAMS:
            build_team(roster, team).to_excel(xw, sheet_name=team, index=False)
        validation.to_excel(xw, sheet_name="Validation", index=False)
        # notes below the validation table
        ws = xw.sheets["Validation"]
        start = len(validation) + 3
        ws.cell(row=start, column=1, value="NOTES / FLAGS").font = Font(bold=True)
        for i, note in enumerate(notes, 1):
            ws.cell(row=start + i, column=1, value=f"- {note}")

    # styling pass
    from openpyxl import load_workbook
    wb = load_workbook(OUT)
    for name in wb.sheetnames:
        ws = wb[name]
        _style_sheet(ws)
        headers = [c.value for c in ws[1]]
        _color_surplus(ws, headers)
        if name == "Summary":  # highlight Kerr row
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == MY_TEAM:
                    for c in ws[row]:
                        if c.fill.fgColor.rgb in ("00000000", None):
                            c.fill = KERR_FILL
    wb.save(OUT)

    print(f"wrote {OUT.relative_to(PROCESSED_DIR.parent.parent)}")
    print(f"  Summary + Targets + Cut_Plan + FA_Pool + {len(TEAMS)} team tabs + Validation")
    print("\n=== Targets (bid plan) ===")
    print(bid_plan[["pri", "player", "pos", "model_fair", "your_max",
                    "premium_vs_model"]].to_string(index=False))
    print("\n=== Cut plan ===")
    print(cut_plan[["pri", "player", "pos", "salary_2026", "surplus_2026",
                    "cap_freed_cut50"]].to_string(index=False))
    print("\n=== Budget (Kerr) ===")
    print(budget.to_string(index=False))


if __name__ == "__main__":
    main()
